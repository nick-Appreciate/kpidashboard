"""
Email Parser for Guest Card Inquiry Reports - Supabase Version
Monitors email inbox for Excel attachments and sends data to Vercel API
"""

import imaplib
import email
from email.header import decode_header
import openpyxl
import csv
from datetime import datetime
import requests
import os
import time
import logging
from io import BytesIO, StringIO
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Configuration
EMAIL_HOST = os.getenv('EMAIL_HOST', 'imap.gmail.com')
EMAIL_PORT = int(os.getenv('EMAIL_PORT', 993))
EMAIL_USER = os.getenv('EMAIL_USER')
EMAIL_PASSWORD = os.getenv('EMAIL_PASSWORD')
API_URL = os.getenv('API_URL', 'http://localhost:3000/api')  # Your Vercel URL
CHECK_INTERVAL = int(os.getenv('CHECK_INTERVAL', 300))  # seconds
SEARCH_SUBJECT = os.getenv('SEARCH_SUBJECT', 'Guest Card Inquiries')

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('parser.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class EmailParser:
    def __init__(self):
        self.mail = None
        self.processed_uids = set()
        
    def connect(self):
        """Connect to email server"""
        try:
            self.mail = imaplib.IMAP4_SSL(EMAIL_HOST, EMAIL_PORT)
            self.mail.login(EMAIL_USER, EMAIL_PASSWORD)
            logger.info(f"Successfully connected to {EMAIL_HOST}")
            return True
        except Exception as e:
            logger.error(f"Failed to connect to email: {e}")
            return False
    
    def disconnect(self):
        """Disconnect from email server"""
        if self.mail:
            try:
                self.mail.close()
                self.mail.logout()
                logger.info("Disconnected from email server")
            except:
                pass
    
    def parse_leasing_report(self, file_content):
        """Parse Leasing Report Excel file and extract inquiry data"""
        try:
            wb = openpyxl.load_workbook(BytesIO(file_content))
            sheet = wb.active

            all_data = list(sheet.iter_rows(values_only=True))
            inquiries = []
            current_property = None
            
            # Find header row and column indices dynamically
            header_row = None
            column_map = {}
            for i, row in enumerate(all_data):
                if row and 'Name' in str(row[0] or '') or (len(row) > 1 and 'Email' in str(row[1] or '')):
                    header_row = i
                    for j, cell in enumerate(row):
                        if cell:
                            column_map[str(cell).strip()] = j
                    break
            
            # Get Inquiry ID column index if it exists
            inquiry_id_col = column_map.get('Inquiry ID')
            logger.info(f"Column map: {column_map}")
            if inquiry_id_col is not None:
                logger.info(f"Found Inquiry ID column at index {inquiry_id_col}")

            for i, row in enumerate(all_data):
                if i < 12:
                    continue

                if row[0] and all(cell is None for cell in row[1:]):
                    current_property = str(row[0])
                elif row[0] and isinstance(row[3], datetime):
                    inquiry = {
                        'property': current_property,
                        'name': str(row[0]) if row[0] else None,
                        'email': str(row[1]) if row[1] else None,
                        'phone': str(row[2]) if row[2] else None,
                        'inquiry_received': row[3].isoformat() if isinstance(row[3], datetime) else str(row[3]),
                        'first_contact': row[4].strftime('%Y-%m-%d') if isinstance(row[4], datetime) else str(row[4]) if row[4] else None,
                        'last_activity_date': row[5].strftime('%Y-%m-%d') if isinstance(row[5], datetime) else str(row[5]) if row[5] else None,
                        'last_activity_type': str(row[6]) if row[6] else None,
                        'status': str(row[7]) if row[7] else None,
                        'move_in_preference': str(row[8]) if row[8] else None,
                        'max_rent': str(row[9]) if row[9] else None,
                        'bed_bath_preference': str(row[10]) if row[10] else None,
                        'pet_preference': str(row[11]) if row[11] else None,
                        'monthly_income': str(row[12]) if row[12] else None,
                        'credit_score': str(row[13]) if row[13] else None,
                        'lead_type': str(row[14]) if row[14] else None,
                        'inquiry_id': str(row[inquiry_id_col]) if inquiry_id_col is not None and len(row) > inquiry_id_col and row[inquiry_id_col] else None
                    }
                    inquiries.append(inquiry)

            logger.info(f"Parsed {len(inquiries)} leasing inquiries from Excel file")
            return inquiries

        except Exception as e:
            logger.error(f"Error parsing Leasing Report Excel file: {e}")
            return None

    def parse_leasing_report_csv(self, file_content):
        """Parse Leasing Report CSV file and extract inquiry data"""
        try:
            # Decode bytes to string
            content_str = file_content.decode('utf-8-sig')  # utf-8-sig handles BOM
            reader = csv.DictReader(StringIO(content_str))
            
            inquiries = []
            current_property = None
            
            # Get field names to check for Inquiry ID column
            fieldnames = reader.fieldnames or []
            logger.info(f"CSV columns: {fieldnames}")
            inquiry_id_field = None
            for field in fieldnames:
                if 'inquiry' in field.lower() and 'id' in field.lower():
                    inquiry_id_field = field
                    logger.info(f"Found Inquiry ID column: {inquiry_id_field}")
                    break
            
            for row in reader:
                name = row.get('Name', '').strip()
                
                # Skip empty rows
                if not name:
                    continue
                
                # Check if this is a property header row (starts with "->")
                if name.startswith('->'):
                    current_property = name.replace('->', '').strip()
                    continue
                
                # Skip if no property context
                if not current_property:
                    continue
                
                # Parse inquiry received date
                inquiry_received_str = row.get('Inquiry Received', '')
                if not inquiry_received_str:
                    continue
                
                # Convert date format "12/29/2025 at 01:19 PM" to ISO format
                try:
                    dt = datetime.strptime(inquiry_received_str, '%m/%d/%Y at %I:%M %p')
                    inquiry_received = dt.isoformat()
                except ValueError:
                    inquiry_received = inquiry_received_str
                
                # Parse first contact date
                first_contact_str = row.get('First Contact Date', '')
                first_contact = None
                if first_contact_str:
                    try:
                        dt = datetime.strptime(first_contact_str, '%m/%d/%Y at %I:%M %p')
                        first_contact = dt.strftime('%Y-%m-%d')
                    except ValueError:
                        try:
                            dt = datetime.strptime(first_contact_str, '%m/%d/%Y')
                            first_contact = dt.strftime('%Y-%m-%d')
                        except ValueError:
                            first_contact = first_contact_str
                
                # Parse last activity date
                last_activity_str = row.get('Last Activity Date', '')
                last_activity_date = None
                if last_activity_str:
                    try:
                        dt = datetime.strptime(last_activity_str, '%m/%d/%Y')
                        last_activity_date = dt.strftime('%Y-%m-%d')
                    except ValueError:
                        last_activity_date = last_activity_str
                
                inquiry = {
                    'property': current_property,
                    'name': name,
                    'email': row.get('Email Address', '').strip() or None,
                    'phone': row.get('Phone Number', '').strip() or None,
                    'inquiry_received': inquiry_received,
                    'first_contact': first_contact,
                    'last_activity_date': last_activity_date,
                    'last_activity_type': row.get('Last Activity Type', '').strip() or None,
                    'status': row.get('Status', '').strip() or None,
                    'move_in_preference': row.get('Move In Preference', '').strip() or None,
                    'max_rent': row.get('Max Rent', '').strip() or None,
                    'bed_bath_preference': row.get('Bed Bath Preference', '').strip() or None,
                    'pet_preference': row.get('Pet Preference', '').strip() or None,
                    'monthly_income': row.get('Monthly Income', '').strip() or None,
                    'credit_score': row.get('Credit Score', '').strip() or None,
                    'lead_type': row.get('Lead Type', '').strip() or None,
                    'source': row.get('Source', '').strip() or None,
                    'unit': row.get('Unit', '').strip() or None,
                    'touch_points': int(row.get('Touch Points', 0) or 0) if row.get('Touch Points', '').strip() else None,
                    'follow_ups': int(row.get('Follow Ups', 0) or 0) if row.get('Follow Ups', '').strip() else None,
                    'inquiry_id': row.get(inquiry_id_field, '').strip() if inquiry_id_field else None
                }
                inquiries.append(inquiry)
            
            logger.info(f"Parsed {len(inquiries)} leasing inquiries from CSV file")
            return inquiries
        
        except Exception as e:
            logger.error(f"Error parsing Leasing Report CSV file: {e}")
            return None

    def parse_property_report(self, file_content):
        """Parse Property Report Excel file and extract unit data"""
        try:
            wb = openpyxl.load_workbook(BytesIO(file_content))
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))

            header_index = None
            for i, row in enumerate(rows):
                if row and len(row) > 1 and row[0] == 'Unit' and row[1] == 'BD/BA':
                    header_index = i
                    break

            if header_index is None:
                logger.error("Property Report header row not found")
                return None

            units = []
            current_property = None

            for row in rows[header_index + 1:]:
                if not row or not any(cell is not None for cell in row):
                    continue

                if row[0] and all(cell is None for cell in row[1:]):
                    current_property = str(row[0])
                    continue

                if not current_property:
                    continue

                unit = row[0]
                if unit is None:
                    continue

                unit_str = str(unit).strip()
                if 'Units' in unit_str and row[1] is None:
                    continue

                def parse_date(value):
                    if isinstance(value, datetime):
                        return value.strftime('%Y-%m-%d')
                    return None

                def parse_number(value):
                    if value is None or value == '':
                        return None
                    return float(value)

                unit_row = {
                    'property': current_property,
                    'unit': unit_str,
                    'bd_ba': str(row[1]) if row[1] else None,
                    'status': str(row[2]) if row[2] else None,
                    'sqft': int(row[3]) if isinstance(row[3], (int, float)) else None,
                    'total': parse_number(row[4]),
                    'past_due': parse_number(row[5]),
                    'other_charges': parse_number(row[6]),
                    'tenant_reimbursement_utilities': parse_number(row[7]),
                    'tenant_rental_income': parse_number(row[8]),
                    'cha_affordable_housing_income': parse_number(row[9]),
                    'iha_affordable_housing_income': parse_number(row[10]),
                    'kckha_affordable_housing_income': parse_number(row[11]),
                    'hakc_affordable_housing_income': parse_number(row[12]),
                    'hud_affordable_housing_income': parse_number(row[13]),
                    'pet_rent': parse_number(row[14]),
                    'storage_fee': parse_number(row[15]),
                    'parking_fee': parse_number(row[16]),
                    'insurance_services': parse_number(row[17]),
                    'lease_from': parse_date(row[18]),
                    'lease_to': parse_date(row[19])
                }
                units.append(unit_row)

            logger.info(f"Parsed {len(units)} property report units from Excel file")
            return units

        except Exception as e:
            logger.error(f"Error parsing Property Report Excel file: {e}")
            return None
    
    def send_to_api(self, endpoint, records, metadata=None):
        """Send parsed data to Vercel API"""
        try:
            payload = {
                'records': records,
                'metadata': metadata or {},
            }
            
            response = requests.post(
                f"{API_URL}/{endpoint}",
                json=payload,
                timeout=30
            )
            
            if response.status_code == 200:
                logger.info(f"Successfully sent {len(records)} records to API")
                return True
            else:
                logger.error(f"API returned error: {response.status_code} - {response.text}")
                return False
                
        except Exception as e:
            logger.error(f"Error sending data to API: {e}")
            return False
    
    def process_email(self, email_id):
        """Process a single email"""
        try:
            # Fetch the email
            status, msg_data = self.mail.fetch(email_id, '(RFC822)')
            
            for response_part in msg_data:
                if isinstance(response_part, tuple):
                    msg = email.message_from_bytes(response_part[1])
                    
                    # Get subject
                    subject = decode_header(msg["Subject"])[0][0]
                    if isinstance(subject, bytes):
                        subject = subject.decode()
                    
                    # Get date
                    date_str = msg["Date"]
                    
                    logger.info(f"Processing email: {subject}")
                    
                    # Look for Excel attachments
                    for part in msg.walk():
                        if part.get_content_maintype() == 'multipart':
                            continue
                        
                        filename = part.get_filename()
                        if filename and filename.endswith(('.xlsx', '.xls')):
                            logger.info(f"Found Excel attachment: {filename}")
                            
                            # Get file content
                            file_content = part.get_payload(decode=True)
                            
                            filename_lower = filename.lower()
                            metadata = {
                                'filename': filename,
                                'email_subject': subject,
                                'email_date': date_str,
                                'received_at': datetime.now().isoformat()
                            }

                            if 'guest_card' in filename_lower or 'leasing' in filename_lower:
                                # Use CSV parser for .csv files, Excel parser for .xlsx
                                if filename_lower.endswith('.csv'):
                                    records = self.parse_leasing_report_csv(file_content)
                                else:
                                    records = self.parse_leasing_report(file_content)
                                endpoint = 'inquiries'
                            elif 'rent_roll' in filename_lower or 'property' in filename_lower:
                                records = self.parse_property_report(file_content)
                                endpoint = 'property-reports'
                            else:
                                logger.error(f"Unknown report type for {filename}")
                                continue

                            if records:
                                if self.send_to_api(endpoint, records, metadata):
                                    logger.info(f"Successfully processed {filename}")
                                    return True
                                logger.error(f"Failed to send data from {filename}")
                            else:
                                logger.error(f"Failed to parse {filename}")
            
            return False
            
        except Exception as e:
            logger.error(f"Error processing email {email_id}: {e}")
            return False
    
    def check_inbox(self):
        """Check inbox for new emails"""
        try:
            # Select inbox
            self.mail.select('INBOX')
            
            # Search for emails with subject containing keyword
            if SEARCH_SUBJECT:
                status, messages = self.mail.search(None, f'SUBJECT "{SEARCH_SUBJECT}"', 'UNSEEN')
            else:
                status, messages = self.mail.search(None, 'UNSEEN')
            
            email_ids = messages[0].split()
            
            if email_ids:
                logger.info(f"Found {len(email_ids)} unread emails")
                
                for email_id in email_ids:
                    email_uid = email_id.decode()
                    
                    # Skip if already processed
                    if email_uid in self.processed_uids:
                        continue
                    
                    # Process the email
                    if self.process_email(email_id):
                        self.processed_uids.add(email_uid)
                        # Mark as read
                        self.mail.store(email_id, '+FLAGS', '\\Seen')
            else:
                logger.debug("No new emails found")
                
        except Exception as e:
            logger.error(f"Error checking inbox: {e}")
            # Try to reconnect
            self.disconnect()
            time.sleep(5)
            self.connect()
    
    def run(self):
        """Main loop to continuously check emails"""
        logger.info("Starting email parser service...")
        logger.info(f"API URL: {API_URL}")
        
        if not self.connect():
            logger.error("Failed to connect to email. Exiting.")
            return
        
        try:
            while True:
                self.check_inbox()
                logger.debug(f"Sleeping for {CHECK_INTERVAL} seconds...")
                time.sleep(CHECK_INTERVAL)
                
        except KeyboardInterrupt:
            logger.info("Shutting down email parser...")
        finally:
            self.disconnect()


if __name__ == "__main__":
    # Validate configuration
    if not EMAIL_USER or not EMAIL_PASSWORD:
        logger.error("EMAIL_USER and EMAIL_PASSWORD must be set in .env file")
        exit(1)
    
    if not API_URL:
        logger.error("API_URL must be set in .env file (your Vercel URL)")
        exit(1)
    
    parser = EmailParser()
    parser.run()
